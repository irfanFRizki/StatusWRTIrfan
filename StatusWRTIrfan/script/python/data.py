#!/usr/bin/env python3
"""
Telegram Bot untuk Manajemen File di OpenWrt
Instalasi:
1. opkg update
2. opkg install python3 python3-pip
3. pip3 install python-telegram-bot openpyxl
4. Edit ADMIN_CHAT_ID dan BOT_TOKEN
5. Jalankan: python3 bot.py
"""

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
import os
import logging
from datetime import datetime
import json
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============ KONFIGURASI ============
BOT_TOKEN = "YOUR_BOT_TOKEN_HERE"  # Token dari @BotFather
ADMIN_CHAT_ID = "YOUR_TELEGRAM_ID_HERE"  # ID Telegram admin (gunakan @userinfobot)
BASE_PATH = "/mnt/sda1/data/"
LOG_FILE = "/mnt/sda1/data/bot_activity.log"
BACKUP_PATH = "/mnt/sda1/data/backup/"  # Folder untuk backup log
PASSWORD = "password123"  # Ganti dengan password yang aman
AUTO_DELETE_SECONDS = 30  # Hapus pesan sensitif setelah 30 detik
AUTO_BACKUP_HOURS = 24  # Backup otomatis setiap 24 jam
MAX_LOG_SIZE_MB = 10  # Ukuran maksimal log sebelum backup otomatis

# ============ STRUKTUR DATA ============
PERSONS = ["IRFAN", "ANITA", "AQILLA", "BABA", "MAMAH"]
DOCUMENT_TYPES = ["KTP", "KK", "AKTE LAHIR", "IJASAH TERAKHIR", "CV"]
FILE_FORMATS = ["JPG", "PDF", "PNG", "DOCX"]

# File untuk menyimpan data dinamis
DATA_FILE = "/mnt/sda1/data/bot_data.json"

# ============ FUNGSI DATA MANAGEMENT ============
def load_data():
    """Load data dari file"""
    global PERSONS, DOCUMENT_TYPES
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                PERSONS = data.get('persons', PERSONS)
                DOCUMENT_TYPES = data.get('document_types', DOCUMENT_TYPES)
                logger.info("Data loaded from file")
    except Exception as e:
        logger.error(f"Error loading data: {e}")

def save_data():
    """Simpan data ke file"""
    try:
        os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
        data = {
            'persons': PERSONS,
            'document_types': DOCUMENT_TYPES
        }
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info("Data saved to file")
    except Exception as e:
        logger.error(f"Error saving data: {e}")

# Setup logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Simpan state user (untuk password verification)
user_states = {}

# Waktu backup terakhir
last_backup_time = None

# ============ FUNGSI BACKUP & EXPORT ============
def create_backup():
    """Buat backup file log"""
    try:
        if not os.path.exists(LOG_FILE):
            logger.info("No log file to backup")
            return None
        
        # Buat folder backup jika belum ada
        os.makedirs(BACKUP_PATH, exist_ok=True)
        
        # Nama file backup dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"bot_activity_backup_{timestamp}.log"
        backup_filepath = os.path.join(BACKUP_PATH, backup_filename)
        
        # Copy file log ke backup
        shutil.copy2(LOG_FILE, backup_filepath)
        
        logger.info(f"Backup created: {backup_filepath}")
        return backup_filepath
    except Exception as e:
        logger.error(f"Error creating backup: {e}")
        return None

def check_auto_backup():
    """Cek apakah perlu backup otomatis"""
    global last_backup_time
    
    try:
        # Cek ukuran file log
        if os.path.exists(LOG_FILE):
            file_size_mb = os.path.getsize(LOG_FILE) / (1024 * 1024)
            
            # Backup jika file terlalu besar
            if file_size_mb >= MAX_LOG_SIZE_MB:
                logger.info(f"Log file size {file_size_mb:.2f}MB, creating backup...")
                backup_path = create_backup()
                if backup_path:
                    # Clear log file setelah backup
                    open(LOG_FILE, 'w').close()
                    last_backup_time = datetime.now()
                    return True, "size_limit"
        
        # Cek waktu backup terakhir
        if last_backup_time:
            hours_since_backup = (datetime.now() - last_backup_time).total_seconds() / 3600
            if hours_since_backup >= AUTO_BACKUP_HOURS:
                logger.info(f"Auto backup triggered ({hours_since_backup:.1f} hours since last backup)")
                create_backup()
                last_backup_time = datetime.now()
                return True, "time_based"
        else:
            # Set initial backup time
            last_backup_time = datetime.now()
        
        return False, None
    except Exception as e:
        logger.error(f"Error in auto backup check: {e}")
        return False, None

def export_log_to_excel():
    """Export log ke file Excel"""
    try:
        if not os.path.exists(LOG_FILE):
            return None
        
        # Baca log file
        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Buat workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Bot Activity Log"
        
        # Header
        headers = ["Timestamp", "User", "Action", "Status", "Details"]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Tambahkan data
        for line in lines:
            try:
                entry = json.loads(line.strip())
                ws.append([
                    entry.get('timestamp', ''),
                    entry.get('user', ''),
                    entry.get('action', ''),
                    entry.get('status', ''),
                    entry.get('details', '')
                ])
            except:
                continue
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Style status cells dengan warna
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                if cell.value == "SUCCESS":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif cell.value == "ERROR":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
                elif cell.value == "WARNING":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C6500")
        
        # Simpan file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"bot_activity_log_{timestamp}.xlsx"
        excel_filepath = os.path.join(BACKUP_PATH, excel_filename)
        
        os.makedirs(BACKUP_PATH, exist_ok=True)
        wb.save(excel_filepath)
        
        logger.info(f"Excel export created: {excel_filepath}")
        return excel_filepath
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        return None

# ============ FUNGSI LOGGING ============
def log_activity(user_info: str, action: str, status: str = "INFO", details: str = ""):
    """Log aktivitas ke file"""
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = {
            "timestamp": timestamp,
            "user": user_info,
            "action": action,
            "status": status,
            "details": details
        }
        
        # Buat folder log jika belum ada
        log_dir = os.path.dirname(LOG_FILE)
        if not os.path.exists(log_dir):
            os.makedirs(log_dir, exist_ok=True)
        
        # Tulis ke file log
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(json.dumps(log_entry, ensure_ascii=False) + "\n")
        
        logger.info(f"Logged: {action} - {user_info}")
        
        # Cek auto backup
        check_auto_backup()
        
    except Exception as e:
        logger.error(f"Error writing log: {e}")

# ============ FUNGSI HELPER ============
async def send_admin_notification(context: ContextTypes.DEFAULT_TYPE, user_info: str, action: str, status: str = "INFO"):
    """Kirim notifikasi ke admin"""
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Tentukan emoji berdasarkan status
        emoji = "🔔"
        if status == "SUCCESS":
            emoji = "✅"
        elif status == "ERROR":
            emoji = "❌"
        elif status == "WARNING":
            emoji = "⚠️"
        
        message = f"{emoji} <b>Aktivitas Bot</b>\n\n"
        message += f"👤 User: {user_info}\n"
        message += f"📋 Aksi: {action}\n"
        message += f"⏰ Waktu: {timestamp}"
        
        await context.bot.send_message(
            chat_id=ADMIN_CHAT_ID,
            text=message,
            parse_mode='HTML'
        )
        
        # Log juga ke file
        log_activity(user_info, action, status)
    except Exception as e:
        logger.error(f"Error sending admin notification: {e}")

def get_user_info(update: Update) -> str:
    """Dapatkan informasi user"""
    user = update.effective_user
    username = f"@{user.username}" if user.username else "No username"
    return f"{user.full_name} ({username}) [ID: {user.id}]"

def find_file(person: str, doc_type: str, file_format: str) -> str:
    """Cari file berdasarkan person, document type, dan format"""
    doc_type_clean = doc_type.replace(" ", "_")
    filename = f"{person}_{doc_type_clean}.{file_format.lower()}"
    filepath = os.path.join(BASE_PATH, filename)
    
    # Coba berbagai variasi ekstensi
    variations = [
        filepath,
        filepath.replace(f".{file_format.lower()}", f".{file_format.upper()}"),
        filepath.replace(f".{file_format.lower()}", f".{file_format.capitalize()}")
    ]
    
    for path in variations:
        if os.path.exists(path):
            return path
    
    return None

async def auto_delete_message(context: ContextTypes.DEFAULT_TYPE, chat_id: int, message_id: int, delay: int):
    """Hapus pesan secara otomatis setelah delay tertentu"""
    try:
        import asyncio
        await asyncio.sleep(delay)
        await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
        logger.info(f"Auto-deleted message {message_id} from chat {chat_id}")
    except Exception as e:
        logger.error(f"Error auto-deleting message: {e}")

# ============ HANDLERS ============
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk command /start"""
    user_id = update.effective_user.id
    user_info = get_user_info(update)
    
    # Cek apakah admin
    is_admin = str(user_id) == str(ADMIN_CHAT_ID)
    
    await send_admin_notification(context, user_info, "Memulai bot dengan /start")
    
    keyboard = []
    for person in PERSONS:
        keyboard.append([InlineKeyboardButton(f"👤 {person}", callback_data=f"person_{person}")])
    
    # Tambah menu admin jika user adalah admin
    if is_admin:
        keyboard.append([InlineKeyboardButton("⚙️ ADMIN MENU", callback_data="admin_menu")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    welcome_text = (
        "🔐 <b>Selamat Datang di File Manager Bot</b>\n\n"
        "Pilih nama untuk mengakses dokumen:"
    )
    
    if is_admin:
        welcome_text += "\n\n👑 <i>Anda login sebagai Admin</i>"
    
    await update.message.reply_text(
        welcome_text,
        reply_markup=reply_markup,
        parse_mode='HTML'
    )

async def admin_menu(query, context):
    """Tampilkan menu admin"""
    keyboard = [
        [InlineKeyboardButton("📤 Upload File", callback_data="admin_upload")],
        [InlineKeyboardButton("➕ Tambah Nama", callback_data="admin_addperson")],
        [InlineKeyboardButton("➕ Tambah Jenis Dokumen", callback_data="admin_adddoctype")],
        [InlineKeyboardButton("📋 Lihat Data", callback_data="admin_viewdata")],
        [InlineKeyboardButton("📊 Lihat Log", callback_data="admin_viewlog")],
        [InlineKeyboardButton("📥 Export Log (Excel)", callback_data="admin_exportlog")],
        [InlineKeyboardButton("💾 Backup Log Sekarang", callback_data="admin_backuplog")],
        [InlineKeyboardButton("📂 Lihat Backup", callback_data="admin_viewbackup")],
        [InlineKeyboardButton("🗑️ Hapus Log", callback_data="admin_clearlog")],
        [InlineKeyboardButton("◀️ Kembali", callback_data="back_main")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    # Info backup terakhir
    backup_info = ""
    if last_backup_time:
        hours_ago = (datetime.now() - last_backup_time).total_seconds() / 3600
        backup_info = f"\n\n📅 Backup terakhir: {hours_ago:.1f} jam yang lalu"
    
    # Info ukuran log
    log_size_info = ""
    if os.path.exists(LOG_FILE):
        file_size_mb = os.path.getsize(LOG_FILE) / (1024 * 1024)
        log_size_info = f"\n📊 Ukuran log: {file_size_mb:.2f} MB"
    
    await query.edit_message_text(
        f"⚙️ <b>ADMIN MENU</b>{backup_info}{log_size_info}\n\nPilih menu admin:",
        reply_markup=reply_markup,
        parse_mode='HTML'
    )

async def admin_upload_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mulai proses upload file"""
    query = update.callback_query
    
    keyboard = []
    for person in PERSONS:
        keyboard.append([InlineKeyboardButton(f"👤 {person}", callback_data=f"upload_person_{person}")])
    keyboard.append([InlineKeyboardButton("◀️ Kembali", callback_data="admin_menu")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(
        "📤 <b>Upload File</b>\n\nPilih nama untuk upload dokumen:",
        reply_markup=reply_markup,
        parse_mode='HTML'
    )

async def admin_view_log(query, context):
    """Tampilkan log aktivitas terakhir"""
    try:
        if not os.path.exists(LOG_FILE):
            await query.edit_message_text(
                "📊 <b>Log Aktivitas</b>\n\n"
                "Belum ada log aktivitas.\n\n"
                "Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
            return
        
        # Baca 20 log terakhir
        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        last_logs = lines[-20:] if len(lines) > 20 else lines
        
        log_text = "📊 <b>Log Aktivitas (20 Terakhir)</b>\n\n"
        
        for line in reversed(last_logs):
            try:
                entry = json.loads(line.strip())
                emoji = "🔔"
                if entry.get("status") == "SUCCESS":
                    emoji = "✅"
                elif entry.get("status") == "ERROR":
                    emoji = "❌"
                elif entry.get("status") == "WARNING":
                    emoji = "⚠️"
                
                log_text += f"{emoji} <code>{entry['timestamp']}</code>\n"
                log_text += f"   {entry['action']}\n"
                log_text += f"   User: {entry['user'][:30]}...\n\n"
            except:
                continue
        
        # Split jika terlalu panjang
        if len(log_text) > 4000:
            log_text = log_text[:4000] + "\n\n<i>... log dipotong ...</i>"
        
        keyboard = [
            [InlineKeyboardButton("🔄 Refresh", callback_data="admin_viewlog")],
            [InlineKeyboardButton("◀️ Kembali", callback_data="admin_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            log_text,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    except Exception as e:
        logger.error(f"Error viewing log: {e}")
        await query.edit_message_text(
            f"❌ Error membaca log: {str(e)}\n\n"
            "Gunakan /start untuk kembali.",
            parse_mode='HTML'
        )

async def admin_clear_log(query, context):
    """Hapus log file"""
    try:
        if os.path.exists(LOG_FILE):
            # Backup dulu sebelum hapus
            backup_path = create_backup()
            
            os.remove(LOG_FILE)
            
            msg = "✅ Log berhasil dihapus!"
            if backup_path:
                msg += f"\n\n💾 Backup tersimpan di:\n<code>{backup_path}</code>"
            
            await query.edit_message_text(
                msg + "\n\nGunakan /start untuk kembali ke menu utama.",
                parse_mode='HTML'
            )
        else:
            await query.edit_message_text(
                "ℹ️ Log sudah kosong.\n\n"
                "Gunakan /start untuk kembali ke menu utama."
            )
    except Exception as e:
        await query.edit_message_text(
            f"❌ Error menghapus log: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def admin_backup_log(query, context):
    """Backup log file sekarang"""
    try:
        status_msg = await query.edit_message_text("⏳ Sedang membuat backup...")
        
        backup_path = create_backup()
        
        if backup_path:
            await status_msg.edit_text(
                f"✅ <b>Backup Berhasil!</b>\n\n"
                f"📁 File: <code>{os.path.basename(backup_path)}</code>\n"
                f"📂 Path: <code>{backup_path}</code>\n\n"
                f"Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
        else:
            await status_msg.edit_text(
                "❌ Gagal membuat backup!\n\n"
                "Gunakan /start untuk kembali."
            )
    except Exception as e:
        await query.edit_message_text(
            f"❌ Error: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def admin_view_data(query, context):
    """Tampilkan data nama dan jenis dokumen"""
    try:
        data_text = "📋 <b>Data Bot</b>\n\n"
        
        data_text += f"👥 <b>Nama ({len(PERSONS)}):</b>\n"
        for i, person in enumerate(PERSONS, 1):
            data_text += f"{i}. {person}\n"
        
        data_text += f"\n📄 <b>Jenis Dokumen ({len(DOCUMENT_TYPES)}):</b>\n"
        for i, doc_type in enumerate(DOCUMENT_TYPES, 1):
            data_text += f"{i}. {doc_type}\n"
        
        data_text += f"\n📎 <b>Format File ({len(FILE_FORMATS)}):</b>\n"
        data_text += f"{', '.join(FILE_FORMATS)}"
        
        keyboard = [
            [InlineKeyboardButton("🔄 Refresh", callback_data="admin_viewdata")],
            [InlineKeyboardButton("◀️ Kembali", callback_data="admin_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            data_text,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    except Exception as e:
        logger.error(f"Error viewing data: {e}")
        await query.edit_message_text(
            f"❌ Error: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def admin_add_person(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mulai proses tambah nama baru"""
    if update.callback_query:
        query = update.callback_query
        context.user_data['awaiting_new_person'] = True
        
        await query.edit_message_text(
            "➕ <b>Tambah Nama Baru</b>\n\n"
            "Silakan ketik nama baru (huruf kapital).\n"
            "Contoh: BUDI\n\n"
            "<i>Ketik /cancel untuk batal</i>",
            parse_mode='HTML'
        )

async def admin_add_doctype(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mulai proses tambah jenis dokumen baru"""
    if update.callback_query:
        query = update.callback_query
        context.user_data['awaiting_new_doctype'] = True
        
        await query.edit_message_text(
            "➕ <b>Tambah Jenis Dokumen Baru</b>\n\n"
            "Silakan ketik jenis dokumen baru (huruf kapital).\n"
            "Contoh: SERTIFIKAT\n\n"
            "<i>Ketik /cancel untuk batal</i>",
            parse_mode='HTML'
        )

async def admin_export_log(query, context, user_info):
    """Export log ke Excel"""
    try:
        status_msg = await query.edit_message_text("⏳ Sedang mengexport ke Excel...")
        
        excel_path = export_log_to_excel()
        
        if excel_path and os.path.exists(excel_path):
            # Kirim file Excel
            with open(excel_path, 'rb') as f:
                await context.bot.send_document(
                    chat_id=query.message.chat_id,
                    document=f,
                    filename=os.path.basename(excel_path),
                    caption=f"📊 Export Log Bot Activity\n📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                )
            
            await status_msg.edit_text(
                f"✅ <b>Export Berhasil!</b>\n\n"
                f"📁 File Excel telah dikirim.\n"
                f"📂 Tersimpan di: <code>{excel_path}</code>\n\n"
                f"Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
            
            log_activity(user_info, "Export log ke Excel", "SUCCESS")
        else:
            await status_msg.edit_text(
                "❌ Gagal export ke Excel!\n\n"
                "Gunakan /start untuk kembali."
            )
    except Exception as e:
        logger.error(f"Error exporting log: {e}")
        await query.edit_message_text(
            f"❌ Error: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def admin_view_backup(query, context):
    """Tampilkan daftar file backup"""
    try:
        if not os.path.exists(BACKUP_PATH):
            await query.edit_message_text(
                "📂 <b>Daftar Backup</b>\n\n"
                "Belum ada file backup.\n\n"
                "Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
            return
        
        # Ambil semua file backup
        backup_files = []
        for filename in os.listdir(BACKUP_PATH):
            filepath = os.path.join(BACKUP_PATH, filename)
            if os.path.isfile(filepath):
                file_size = os.path.getsize(filepath) / 1024  # KB
                mod_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                backup_files.append({
                    'name': filename,
                    'size': file_size,
                    'time': mod_time
                })
        
        # Sort by time (newest first)
        backup_files.sort(key=lambda x: x['time'], reverse=True)
        
        if not backup_files:
            await query.edit_message_text(
                "📂 <b>Daftar Backup</b>\n\n"
                "Belum ada file backup.\n\n"
                "Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
            return
        
        # Buat list backup (max 10)
        backup_text = "📂 <b>Daftar Backup</b>\n\n"
        for i, backup in enumerate(backup_files[:10], 1):
            backup_text += f"{i}. <code>{backup['name']}</code>\n"
            backup_text += f"   📊 {backup['size']:.1f} KB | 📅 {backup['time'].strftime('%Y-%m-%d %H:%M')}\n\n"
        
        if len(backup_files) > 10:
            backup_text += f"\n<i>... dan {len(backup_files) - 10} file lainnya</i>\n"
        
        backup_text += f"\n📁 Total: {len(backup_files)} file backup"
        
        keyboard = [
            [InlineKeyboardButton("🔄 Refresh", callback_data="admin_viewbackup")],
            [InlineKeyboardButton("◀️ Kembali", callback_data="admin_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            backup_text,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    except Exception as e:
        logger.error(f"Error viewing backup: {e}")
        await query.edit_message_text(
            f"❌ Error: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk inline keyboard button"""
    query = update.callback_query
    await query.answer()
    
    data = query.data
    user_id = update.effective_user.id
    user_info = get_user_info(update)
    is_admin = str(user_id) == str(ADMIN_CHAT_ID)
    
    # Admin Menu Handlers
    if data == "admin_menu":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_menu(query, context)
        return
    
    elif data == "admin_upload":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_upload_start(update, context)
        return
    
    elif data == "admin_viewlog":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_view_log(query, context)
        return
    
    elif data == "admin_clearlog":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_clear_log(query, context)
        return
    
    elif data == "admin_backuplog":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_backup_log(query, context)
        return
    
    elif data == "admin_exportlog":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_export_log(query, context, user_info)
        return
    
    elif data == "admin_viewbackup":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_view_backup(query, context)
        return
    
    elif data == "admin_viewdata":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_view_data(query, context)
        return
    
    elif data == "admin_addperson":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_add_person(update, context)
        return
    
    elif data == "admin_adddoctype":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_add_doctype(update, context)
        return
    
    # Upload handlers
    elif data.startswith("upload_person_"):
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak!")
            return
        
        person = data.replace("upload_person_", "")
        context.user_data['upload_person'] = person
        
        keyboard = []
        for doc_type in DOCUMENT_TYPES:
            keyboard.append([InlineKeyboardButton(
                f"📄 {doc_type}", 
                callback_data=f"upload_doc_{person}_{doc_type}"
            )])
        keyboard.append([InlineKeyboardButton("◀️ Kembali", callback_data="admin_upload")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            f"📤 <b>Upload untuk {person}</b>\n\nPilih jenis dokumen:",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
        return
    
    elif data.startswith("upload_doc_"):
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak!")
            return
        
        parts = data.replace("upload_doc_", "").split("_", 1)
        person = parts[0]
        doc_type = parts[1]
        
        context.user_data['upload_person'] = person
        context.user_data['upload_doc_type'] = doc_type
        context.user_data['awaiting_file'] = True
        
        await query.edit_message_text(
            f"📤 <b>Upload File</b>\n\n"
            f"👤 Nama: {person}\n"
            f"📄 Dokumen: {doc_type}\n\n"
            f"Silakan kirim file (JPG, PNG, PDF, atau DOCX)\n\n"
            f"<i>File akan disimpan sebagai:\n{person}_{doc_type.replace(' ', '_')}.EXT</i>",
            parse_mode='HTML'
        )
        return
    
    # Handler untuk pilihan person
    elif data.startswith("person_"):
        person = data.replace("person_", "")
        context.user_data['selected_person'] = person
        
        await send_admin_notification(context, user_info, f"Memilih person: {person}")
        
        keyboard = []
        for doc_type in DOCUMENT_TYPES:
            keyboard.append([InlineKeyboardButton(
                f"📄 {doc_type}", 
                callback_data=f"doc_{person}_{doc_type}"
            )])
        keyboard.append([InlineKeyboardButton("◀️ Kembali", callback_data="back_main")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            f"📂 <b>Dokumen untuk {person}</b>\n\nPilih jenis dokumen:",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    
    # Handler untuk pilihan document type
    elif data.startswith("doc_"):
        parts = data.replace("doc_", "").split("_", 1)
        person = parts[0]
        doc_type = parts[1]
        
        context.user_data['selected_doc_type'] = doc_type
        
        await send_admin_notification(context, user_info, f"Memilih dokumen: {person} - {doc_type}")
        
        keyboard = []
        for fmt in FILE_FORMATS:
            keyboard.append([InlineKeyboardButton(
                f"📎 {fmt}", 
                callback_data=f"fmt_{person}_{doc_type}_{fmt}"
            )])
        keyboard.append([InlineKeyboardButton("◀️ Kembali", callback_data=f"person_{person}")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            f"📂 <b>{person} - {doc_type}</b>\n\nPilih format file:",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    
    # Handler untuk pilihan format file
    elif data.startswith("fmt_"):
        parts = data.replace("fmt_", "").split("_")
        person = parts[0]
        doc_type = "_".join(parts[1:-1])
        file_format = parts[-1]
        
        # Cek apakah user sudah terverifikasi
        if user_id not in user_states or not user_states[user_id].get('verified', False):
            context.user_data['pending_request'] = {
                'person': person,
                'doc_type': doc_type,
                'format': file_format
            }
            
            msg = await query.edit_message_text(
                "🔐 <b>Autentikasi Diperlukan</b>\n\n"
                "Silakan masukkan password untuk mengakses file ini:",
                parse_mode='HTML'
            )
            
            user_states[user_id] = {
                'awaiting_password': True,
                'prompt_message_id': msg.message_id
            }
            return
        
        # Jika sudah terverifikasi, kirim file
        await send_file(query, context, person, doc_type, file_format, user_info)
    
    # Handler untuk back to main menu
    elif data == "back_main":
        keyboard = []
        for person in PERSONS:
            keyboard.append([InlineKeyboardButton(f"👤 {person}", callback_data=f"person_{person}")])
        
        if is_admin:
            keyboard.append([InlineKeyboardButton("⚙️ ADMIN MENU", callback_data="admin_menu")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            "🔐 <b>File Manager Bot</b>\n\nPilih nama untuk mengakses dokumen:",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )

async def send_file(query, context, person, doc_type, file_format, user_info):
    """Kirim file ke user"""
    filepath = find_file(person, doc_type, file_format)
    
    if filepath and os.path.exists(filepath):
        try:
            await query.edit_message_text("⏳ Sedang mengirim file...")
            
            # Kirim file berdasarkan tipe
            if file_format.upper() in ["JPG", "PNG"]:
                with open(filepath, 'rb') as f:
                    sent_msg = await context.bot.send_photo(
                        chat_id=query.message.chat_id,
                        photo=f,
                        caption=f"📄 {person} - {doc_type}.{file_format}"
                    )
            else:
                with open(filepath, 'rb') as f:
                    sent_msg = await context.bot.send_document(
                        chat_id=query.message.chat_id,
                        document=f,
                        caption=f"📄 {person} - {doc_type}.{file_format}"
                    )
            
            await send_admin_notification(
                context, 
                user_info, 
                f"Mengunduh file: {person} - {doc_type} - {file_format}",
                "SUCCESS"
            )
            
            success_msg = await query.edit_message_text(
                "✅ File berhasil dikirim!\n\n"
                "Gunakan /start untuk kembali ke menu utama."
            )
            
            # Auto-delete file yang dikirim setelah beberapa detik
            context.application.create_task(
                auto_delete_message(context, query.message.chat_id, sent_msg.message_id, AUTO_DELETE_SECONDS)
            )
            
        except Exception as e:
            logger.error(f"Error sending file: {e}")
            await query.edit_message_text(
                "❌ Terjadi kesalahan saat mengirim file.\n\n"
                "Gunakan /start untuk kembali ke menu utama."
            )
            
            await send_admin_notification(
                context, 
                user_info, 
                f"Error mengirim file: {person} - {doc_type} - {file_format}",
                "ERROR"
            )
    else:
        # Notifikasi ke admin bahwa file tidak ada
        await send_admin_notification(
            context, 
            user_info, 
            f"⚠️ File tidak ditemukan: {person} - {doc_type} - {file_format}",
            "WARNING"
        )
        
        # Kirim notifikasi khusus ke admin tentang file yang hilang
        try:
            doc_type_clean = doc_type.replace(" ", "_")
            expected_filename = f"{person}_{doc_type_clean}.{file_format.lower()}"
            
            await context.bot.send_message(
                chat_id=ADMIN_CHAT_ID,
                text=(
                    f"⚠️ <b>FILE TIDAK DITEMUKAN</b>\n\n"
                    f"👤 User: {user_info}\n"
                    f"📁 File: <code>{expected_filename}</code>\n"
                    f"📂 Path: <code>{BASE_PATH}</code>\n"
                    f"⏰ Waktu: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                    f"<i>Silakan upload file yang diminta atau periksa path file.</i>"
                ),
                parse_mode='HTML'
            )
        except:
            pass
        
        await query.edit_message_text(
            f"❌ File tidak ditemukan!\n\n"
            f"File: {person}_{doc_type}.{file_format}\n"
            f"Path: {BASE_PATH}\n\n"
            f"Admin telah diberitahu tentang file yang hilang.\n\n"
            "Gunakan /start untuk kembali ke menu utama."
        )

async def handle_file_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk upload file dari admin"""
    user_id = update.effective_user.id
    user_info = get_user_info(update)
    
    # Cek apakah admin
    if str(user_id) != str(ADMIN_CHAT_ID):
        await update.message.reply_text("❌ Anda tidak memiliki akses untuk upload file.")
        return
    
    # Cek apakah sedang dalam mode upload
    if not context.user_data.get('awaiting_file', False):
        return
    
    try:
        person = context.user_data.get('upload_person')
        doc_type = context.user_data.get('upload_doc_type')
        
        # Dapatkan file
        file_obj = None
        file_ext = None
        
        if update.message.photo:
            file_obj = update.message.photo[-1]  # Ambil resolusi tertinggi
            file_ext = "JPG"
        elif update.message.document:
            file_obj = update.message.document
            # Dapatkan ekstensi dari nama file
            filename = file_obj.file_name
            file_ext = filename.split('.')[-1].upper()
        
        if not file_obj:
            await update.message.reply_text("❌ Format file tidak didukung!")
            return
        
        # Validasi ekstensi
        if file_ext not in FILE_FORMATS:
            await update.message.reply_text(
                f"❌ Format {file_ext} tidak didukung!\n"
                f"Format yang didukung: {', '.join(FILE_FORMATS)}"
            )
            return
        
        # Download file
        status_msg = await update.message.reply_text("⏳ Sedang mengupload file...")
        
        file = await file_obj.get_file()
        
        # Simpan file dengan nama yang sesuai
        doc_type_clean = doc_type.replace(" ", "_")
        filename = f"{person}_{doc_type_clean}.{file_ext.lower()}"
        filepath = os.path.join(BASE_PATH, filename)
        
        # Buat folder jika belum ada
        os.makedirs(BASE_PATH, exist_ok=True)
        
        await file.download_to_drive(filepath)
        
        # Clear upload state
        context.user_data['awaiting_file'] = False
        context.user_data['upload_person'] = None
        context.user_data['upload_doc_type'] = None
        
        await status_msg.edit_text(
            f"✅ <b>File berhasil diupload!</b>\n\n"
            f"📁 Nama file: <code>{filename}</code>\n"
            f"📂 Path: <code>{filepath}</code>\n"
            f"👤 Person: {person}\n"
            f"📄 Dokumen: {doc_type}\n\n"
            f"Gunakan /start untuk kembali ke menu.",
            parse_mode='HTML'
        )
        
        # Log aktivitas
        log_activity(
            user_info,
            f"Upload file: {filename}",
            "SUCCESS",
            f"Person: {person}, Doc: {doc_type}"
        )
        
        # Delete file asli yang diupload user (untuk keamanan)
        context.application.create_task(
            auto_delete_message(context, update.message.chat_id, update.message.message_id, 5)
        )
        
    except Exception as e:
        logger.error(f"Error uploading file: {e}")
        await update.message.reply_text(
            f"❌ Error saat upload file: {str(e)}\n\n"
            "Gunakan /start untuk coba lagi."
        )
        
        context.user_data['awaiting_file'] = False

async def handle_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk input password dan input teks lainnya"""
    user_id = update.effective_user.id
    user_info = get_user_info(update)
    is_admin = str(user_id) == str(ADMIN_CHAT_ID)
    
    # Skip jika sedang upload file
    if context.user_data.get('awaiting_file', False):
        return
    
    # Handler untuk tambah nama baru (admin only)
    if is_admin and context.user_data.get('awaiting_new_person', False):
        new_person = update.message.text.strip().upper()
        
        # Validasi
        if not new_person:
            await update.message.reply_text("❌ Nama tidak boleh kosong!")
            return
        
        if len(new_person) > 30:
            await update.message.reply_text("❌ Nama terlalu panjang (max 30 karakter)!")
            return
        
        if new_person in PERSONS:
            await update.message.reply_text(f"❌ Nama '{new_person}' sudah ada dalam daftar!")
            return
        
        # Tambahkan nama baru
        PERSONS.append(new_person)
        save_data()
        
        context.user_data['awaiting_new_person'] = False
        
        await update.message.reply_text(
            f"✅ <b>Nama berhasil ditambahkan!</b>\n\n"
            f"👤 Nama: {new_person}\n"
            f"📊 Total nama: {len(PERSONS)}\n\n"
            f"Gunakan /start untuk kembali ke menu.",
            parse_mode='HTML'
        )
        
        log_activity(user_info, f"Menambah nama baru: {new_person}", "SUCCESS")
        return
    
    # Handler untuk tambah jenis dokumen baru (admin only)
    if is_admin and context.user_data.get('awaiting_new_doctype', False):
        new_doctype = update.message.text.strip().upper()
        
        # Validasi
        if not new_doctype:
            await update.message.reply_text("❌ Jenis dokumen tidak boleh kosong!")
            return
        
        if len(new_doctype) > 50:
            await update.message.reply_text("❌ Nama jenis dokumen terlalu panjang (max 50 karakter)!")
            return
        
        if new_doctype in DOCUMENT_TYPES:
            await update.message.reply_text(f"❌ Jenis dokumen '{new_doctype}' sudah ada dalam daftar!")
            return
        
        # Tambahkan jenis dokumen baru
        DOCUMENT_TYPES.append(new_doctype)
        save_data()
        
        context.user_data['awaiting_new_doctype'] = False
        
        await update.message.reply_text(
            f"✅ <b>Jenis dokumen berhasil ditambahkan!</b>\n\n"
            f"📄 Dokumen: {new_doctype}\n"
            f"📊 Total jenis dokumen: {len(DOCUMENT_TYPES)}\n\n"
            f"Gunakan /start untuk kembali ke menu.",
            parse_mode='HTML'
        )
        
        log_activity(user_info, f"Menambah jenis dokumen baru: {new_doctype}", "SUCCESS")
        return
    
    # Handler untuk password
    if user_id in user_states and user_states[user_id].get('awaiting_password', False):
        password_input = update.message.text.strip()
        
        # Hapus pesan password dari user (untuk keamanan)
        context.application.create_task(
            auto_delete_message(context, update.message.chat_id, update.message.message_id, 2)
        )
        
        # Hapus juga pesan prompt password
        if 'prompt_message_id' in user_states[user_id]:
            try:
                context.application.create_task(
                    auto_delete_message(
                        context, 
                        update.message.chat_id, 
                        user_states[user_id]['prompt_message_id'], 
                        2
                    )
                )
            except:
                pass
        
        if password_input == PASSWORD:
            user_states[user_id]['verified'] = True
            user_states[user_id]['awaiting_password'] = False
            
            await send_admin_notification(context, user_info, "Password benar - akses diberikan", "SUCCESS")
            
            # Ambil pending request
            pending = context.user_data.get('pending_request', {})
            if pending:
                person = pending['person']
                doc_type = pending['doc_type']
                file_format = pending['format']
                
                filepath = find_file(person, doc_type, file_format)
                
                if filepath and os.path.exists(filepath):
                    try:
                        status_msg = await update.message.reply_text("⏳ Password benar! Sedang mengirim file...")
                        
                        if file_format.upper() in ["JPG", "PNG"]:
                            with open(filepath, 'rb') as f:
                                sent_msg = await context.bot.send_photo(
                                    chat_id=update.effective_chat.id,
                                    photo=f,
                                    caption=f"📄 {person} - {doc_type}.{file_format}"
                                )
                        else:
                            with open(filepath, 'rb') as f:
                                sent_msg = await context.bot.send_document(
                                    chat_id=update.effective_chat.id,
                                    document=f,
                                    caption=f"📄 {person} - {doc_type}.{file_format}"
                                )
                        
                        await send_admin_notification(
                            context, 
                            user_info, 
                            f"Mengunduh file: {person} - {doc_type} - {file_format}",
                            "SUCCESS"
                        )
                        
                        success_msg = await update.message.reply_text(
                            "✅ File berhasil dikirim!\n\n"
                            f"<i>File akan otomatis terhapus dalam {AUTO_DELETE_SECONDS} detik</i>\n\n"
                            "Gunakan /start untuk kembali ke menu utama.",
                            parse_mode='HTML'
                        )
                        
                        # Auto-delete messages
                        context.application.create_task(
                            auto_delete_message(context, update.effective_chat.id, status_msg.message_id, 5)
                        )
                        context.application.create_task(
                            auto_delete_message(context, update.effective_chat.id, sent_msg.message_id, AUTO_DELETE_SECONDS)
                        )
                        context.application.create_task(
                            auto_delete_message(context, update.effective_chat.id, success_msg.message_id, AUTO_DELETE_SECONDS)
                        )
                        
                    except Exception as e:
                        logger.error(f"Error sending file: {e}")
                        await update.message.reply_text(
                            "❌ Terjadi kesalahan saat mengirim file."
                        )
                else:
                    # Notifikasi ke admin bahwa file tidak ada
                    await send_admin_notification(
                        context, 
                        user_info, 
                        f"⚠️ File tidak ditemukan: {person} - {doc_type} - {file_format}",
                        "WARNING"
                    )
                    
                    # Kirim notifikasi khusus ke admin
                    try:
                        doc_type_clean = doc_type.replace(" ", "_")
                        expected_filename = f"{person}_{doc_type_clean}.{file_format.lower()}"
                        
                        await context.bot.send_message(
                            chat_id=ADMIN_CHAT_ID,
                            text=(
                                f"⚠️ <b>FILE TIDAK DITEMUKAN</b>\n\n"
                                f"👤 User: {user_info}\n"
                                f"📁 File: <code>{expected_filename}</code>\n"
                                f"📂 Path: <code>{BASE_PATH}</code>\n"
                                f"⏰ Waktu: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                                f"<i>Silakan upload file yang diminta atau periksa path file.</i>"
                            ),
                            parse_mode='HTML'
                        )
                    except:
                        pass
                    
                    await update.message.reply_text(
                        f"❌ File tidak ditemukan!\n\n"
                        f"Admin telah diberitahu tentang file yang hilang.\n\n"
                        "Gunakan /start untuk kembali ke menu utama."
                    )
        else:
            await send_admin_notification(context, user_info, "Password salah - akses ditolak", "WARNING")
            
            error_msg = await update.message.reply_text(
                "❌ <b>Password Salah!</b>\n\n"
                "Akses ditolak. Silakan coba lagi dengan /start",
                parse_mode='HTML'
            )
            user_states[user_id]['awaiting_password'] = False
            
            # Auto-delete error message
            context.application.create_task(
                auto_delete_message(context, update.message.chat_id, error_msg.message_id, 10)
            )

async def cancel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk command /cancel"""
    user_id = update.effective_user.id
    
    # Clear semua state
    context.user_data['awaiting_new_person'] = False
    context.user_data['awaiting_new_doctype'] = False
    context.user_data['awaiting_file'] = False
    
    if user_id in user_states:
        user_states[user_id]['awaiting_password'] = False
    
    await update.message.reply_text(
        "❌ <b>Dibatalkan</b>\n\n"
        "Gunakan /start untuk kembali ke menu utama.",
        parse_mode='HTML'
    )

# ============ MAIN ============
def main():
    """Jalankan bot"""
    # Validasi konfigurasi
    if BOT_TOKEN == "YOUR_BOT_TOKEN_HERE" or ADMIN_CHAT_ID == "YOUR_TELEGRAM_ID_HERE":
        print("❌ ERROR: Harap edit BOT_TOKEN dan ADMIN_CHAT_ID terlebih dahulu!")
        return
    
    if not os.path.exists(BASE_PATH):
        print(f"⚠️  WARNING: Folder {BASE_PATH} tidak ditemukan!")
        print(f"Bot akan tetap berjalan, tapi file tidak akan ditemukan.")
    
    # Load data dari file
    load_data()
    
    # Buat aplikasi
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Tambahkan handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("cancel", cancel_command))
    application.add_handler(CallbackQueryHandler(button_callback))
    
    # Handler untuk file upload (harus sebelum text handler)
    application.add_handler(MessageHandler(
        (filters.Document.ALL | filters.PHOTO) & ~filters.COMMAND, 
        handle_file_upload
    ))
    
    # Handler untuk text/password
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_password))
    
    # Jalankan bot
    print("✅ Bot berjalan...")
    print(f"📁 Base Path: {BASE_PATH}")
    print(f"📊 Log File: {LOG_FILE}")
    print(f"💾 Backup Path: {BACKUP_PATH}")
    print(f"💽 Data File: {DATA_FILE}")
    print(f"👤 Admin ID: {ADMIN_CHAT_ID}")
    print(f"🗑️ Auto-delete: {AUTO_DELETE_SECONDS} detik")
    print(f"⏰ Auto-backup: Setiap {AUTO_BACKUP_HOURS} jam atau {MAX_LOG_SIZE_MB}MB")
    print(f"\n📋 Data:")
    print(f"   👥 Nama: {len(PERSONS)} orang")
    print(f"   📄 Jenis Dokumen: {len(DOCUMENT_TYPES)} jenis")
    print(f"\n📋 Fitur:")
    print(f"   ✓ Log aktivitas otomatis")
    print(f"   ✓ Backup otomatis")
    print(f"   ✓ Export ke Excel")
    print(f"   ✓ Notifikasi file tidak ada")
    print(f"   ✓ Auto-delete pesan sensitif")
    print(f"   ✓ Tambah nama dinamis (admin)")
    print(f"   ✓ Tambah jenis dokumen dinamis (admin)")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
