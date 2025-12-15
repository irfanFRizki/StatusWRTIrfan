#!/usr/bin/env python3
"""
Telegram Bot untuk Manajemen File di OpenWrt
Instalasi:
1. opkg update
2. opkg install python3 python3-pip
3. pip3 install python-telegram-bot openpyxl
4. Edit ADMIN_CHAT_ID dan BOT_TOKEN
5. Jalankan: python3 bot.py
"""

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
import os
import logging
from datetime import datetime
import json
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import secrets
import string

# Setup logging PERTAMA sebelum fungsi lain
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ============ KONFIGURASI ============
BOT_TOKEN = "8568041711:AAGInAzY3xI8gJuKX2xCDyvO-pwcWlwLieU"  # Token dari @BotFather
ADMIN_CHAT_ID = "5645537022"  # ID Telegram admin (gunakan @userinfobot)

# Auto-detect BASE_PATH dari mount point yang tersedia
def detect_base_path():
    """Deteksi otomatis BASE_PATH dari /mnt/sd* yang memiliki file dokumen"""
    # File system bot yang harus diabaikan
    bot_files = [
        'bot_activity.log',
        'bot_data.json',
        'bot_service.log',
        '.bot_test',
        '.bot_write_test'
    ]
    
    bot_folders = ['backup', '__pycache__']
    
    # Cari mount point yang tersedia
    mount_points = [
        "/mnt/sda1/data/",
        "/mnt/sdb1/data/",
        "/mnt/sdc1/data/",
        "/mnt/sdd1/data/",
        "/mnt/sde1/data/",
        "/mnt/mmcblk0p1/data/",  # SD Card
        "/mnt/mmcblk1p1/data/",
    ]
    
    # Cek mount point yang memiliki file dokumen
    for path in mount_points:
        if os.path.exists(path):
            try:
                # Cek apakah ada file dengan format NAMA_DOKUMEN.ext
                files = os.listdir(path)
                has_document = False
                
                for file in files:
                    # Skip hidden files, folder, dan file system bot
                    if file.startswith('.') or os.path.isdir(os.path.join(path, file)):
                        continue
                    
                    if file in bot_files:
                        continue
                    
                    if any(folder in file for folder in bot_folders):
                        continue
                    
                    # Cek format file NAMA_DOKUMEN.ext (harus ada underscore)
                    if '_' in file and '.' in file:
                        # Ekstrak nama dan cek apakah valid
                        parts = file.split('_', 1)
                        if len(parts[0]) > 0:  # Ada nama person
                            # Cek extension valid
                            ext = file.split('.')[-1].upper()
                            if ext in ['JPG', 'JPEG', 'PDF', 'PNG', 'DOCX', 'DOC']:
                                has_document = True
                                logger.info(f"Found document: {file} in {path}")
                                break
                
                if has_document:
                    # Test write untuk memastikan bisa digunakan
                    test_file = os.path.join(path, '.bot_test')
                    try:
                        with open(test_file, 'w') as f:
                            f.write('test')
                        os.remove(test_file)
                        logger.info(f"Using BASE_PATH: {path}")
                        return path
                    except:
                        logger.warning(f"Cannot write to {path}")
                        continue
                else:
                    logger.info(f"No valid documents found in {path}")
            except Exception as e:
                logger.warning(f"Cannot access {path}: {e}")
                continue
    
    # Jika tidak ada yang memiliki dokumen, cek mount point yang ada dan writable
    logger.warning("No documents found in any mount point, checking for writable storage...")
    for path in mount_points:
        if os.path.exists(os.path.dirname(path)):
            try:
                os.makedirs(path, exist_ok=True)
                test_file = os.path.join(path, '.bot_test')
                with open(test_file, 'w') as f:
                    f.write('test')
                os.remove(test_file)
                logger.info(f"Using empty but writable BASE_PATH: {path}")
                return path
            except Exception as e:
                logger.warning(f"Cannot use {path}: {e}")
                continue
    
    # Fallback ke /tmp jika tidak ada mount point
    fallback_path = "/tmp/bot_data/"
    logger.warning(f"No external storage found, using fallback: {fallback_path}")
    os.makedirs(fallback_path, exist_ok=True)
    return fallback_path

BASE_PATH = detect_base_path()
LOG_FILE = os.path.join(BASE_PATH, "bot_activity.log")
BACKUP_PATH = os.path.join(BASE_PATH, "backup/")
DATA_FILE = os.path.join(BASE_PATH, "bot_data.json")

# PASSWORD tidak digunakan lagi, diganti dengan approval system
AUTO_DELETE_SECONDS = 30  # Hapus pesan sensitif setelah 30 detik
AUTO_BACKUP_HOURS = 24  # Backup otomatis setiap 24 jam
MAX_LOG_SIZE_MB = 10  # Ukuran maksimal log sebelum backup otomatis

# ============ STRUKTUR DATA ============
PERSONS = ["IRFAN", "ANITA", "AQILLA", "BABA", "MAMAH"]
DOCUMENT_TYPES = ["KTP", "KK", "AKTE LAHIR", "IJASAH TERAKHIR", "CV"]
FILE_FORMATS = ["JPG", "PDF", "PNG", "DOCX"]

# Simpan state user (untuk password verification)
user_states = {}

# Waktu backup terakhir
last_backup_time = None

# Simpan pending approval requests
pending_approvals = {}

# Simpan password yang di-generate untuk setiap user
user_passwords = {}

# ============ FUNGSI STORAGE INFO ============
def get_storage_info():
    """Dapatkan informasi storage yang sedang digunakan"""
    try:
        import subprocess
        
        # Dapatkan info disk usage
        result = subprocess.run(['df', '-h', BASE_PATH], 
                              capture_output=True, 
                              text=True, 
                              timeout=5)
        
        if result.returncode == 0:
            lines = result.stdout.strip().split('\n')
            if len(lines) >= 2:
                # Parse output df
                header = lines[0].split()
                data = lines[1].split()
                
                info = {
                    'device': data[0] if len(data) > 0 else 'Unknown',
                    'size': data[1] if len(data) > 1 else 'Unknown',
                    'used': data[2] if len(data) > 2 else 'Unknown',
                    'available': data[3] if len(data) > 3 else 'Unknown',
                    'use_percent': data[4] if len(data) > 4 else 'Unknown',
                    'mounted': data[5] if len(data) > 5 else 'Unknown',
                    'path': BASE_PATH
                }
                return info
    except Exception as e:
        logger.error(f"Error getting storage info: {e}")
    
    # Fallback info
    return {
        'device': 'Unknown',
        'size': 'Unknown',
        'used': 'Unknown',
        'available': 'Unknown',
        'use_percent': 'Unknown',
        'mounted': 'Unknown',
        'path': BASE_PATH
    }

def get_mount_device():
    """Dapatkan device name dari BASE_PATH (sda1, sdb1, etc)"""
    try:
        # Extract device dari path
        if '/mnt/sda1' in BASE_PATH:
            return 'sda1'
        elif '/mnt/sdb1' in BASE_PATH:
            return 'sdb1'
        elif '/mnt/sdc1' in BASE_PATH:
            return 'sdc1'
        elif '/mnt/sdd1' in BASE_PATH:
            return 'sdd1'
        elif '/mnt/sde1' in BASE_PATH:
            return 'sde1'
        elif '/mnt/mmcblk0p1' in BASE_PATH:
            return 'mmcblk0p1'
        elif '/mnt/mmcblk1p1' in BASE_PATH:
            return 'mmcblk1p1'
        elif '/tmp' in BASE_PATH:
            return 'tmp (Internal)'
        else:
            return 'Unknown'
    except:
        return 'Unknown'

# ============ FUNGSI DATA MANAGEMENT ============
def load_data():
    """Load data dari file"""
    global PERSONS, DOCUMENT_TYPES
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                PERSONS = data.get('persons', PERSONS)
                DOCUMENT_TYPES = data.get('document_types', DOCUMENT_TYPES)
                logger.info("Data loaded from file")
    except Exception as e:
        logger.error(f"Error loading data: {e}")

def save_data():
    """Simpan data ke file"""
    try:
        os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
        data = {
            'persons': PERSONS,
            'document_types': DOCUMENT_TYPES
        }
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info("Data saved to file")
    except Exception as e:
        logger.error(f"Error saving data: {e}")

# ============ FUNGSI PASSWORD ============
def generate_password(length=6):
    """Generate password acak dengan huruf dan angka"""
    # Kombinasi huruf besar, kecil, dan angka
    characters = string.ascii_uppercase + string.digits
    password = ''.join(secrets.choice(characters) for _ in range(length))
    return password

# ============ FUNGSI BACKUP & EXPORT ============
def create_backup():
    """Buat backup file log"""
    try:
        if not os.path.exists(LOG_FILE):
            logger.info("No log file to backup")
            return None
        
        # Buat folder backup jika belum ada
        os.makedirs(BACKUP_PATH, exist_ok=True)
        
        # Nama file backup dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"bot_activity_backup_{timestamp}.log"
        backup_filepath = os.path.join(BACKUP_PATH, backup_filename)
        
        # Copy file log ke backup
        shutil.copy2(LOG_FILE, backup_filepath)
        
        logger.info(f"Backup created: {backup_filepath}")
        return backup_filepath
    except Exception as e:
        logger.error(f"Error creating backup: {e}")
        return None

def check_auto_backup():
    """Cek apakah perlu backup otomatis"""
    global last_backup_time
    
    try:
        # Cek ukuran file log
        if os.path.exists(LOG_FILE):
            file_size_mb = os.path.getsize(LOG_FILE) / (1024 * 1024)
            
            # Backup jika file terlalu besar
            if file_size_mb >= MAX_LOG_SIZE_MB:
                logger.info(f"Log file size {file_size_mb:.2f}MB, creating backup...")
                backup_path = create_backup()
                if backup_path:
                    # Clear log file setelah backup
                    open(LOG_FILE, 'w').close()
                    last_backup_time = datetime.now()
                    return True, "size_limit"
        
        # Cek waktu backup terakhir
        if last_backup_time:
            hours_since_backup = (datetime.now() - last_backup_time).total_seconds() / 3600
            if hours_since_backup >= AUTO_BACKUP_HOURS:
                logger.info(f"Auto backup triggered ({hours_since_backup:.1f} hours since last backup)")
                create_backup()
                last_backup_time = datetime.now()
                return True, "time_based"
        else:
            # Set initial backup time
            last_backup_time = datetime.now()
        
        return False, None
    except Exception as e:
        logger.error(f"Error in auto backup check: {e}")
        return False, None

def export_log_to_excel():
    """Export log ke file Excel"""
    try:
        if not os.path.exists(LOG_FILE):
            return None
        
        # Baca log file
        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Buat workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Bot Activity Log"
        
        # Header
        headers = ["Timestamp", "User", "Action", "Status", "Details"]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Tambahkan data
        for line in lines:
            try:
                entry = json.loads(line.strip())
                ws.append([
                    entry.get('timestamp', ''),
                    entry.get('user', ''),
                    entry.get('action', ''),
                    entry.get('status', ''),
                    entry.get('details', '')
                ])
            except:
                continue
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Style status cells dengan warna
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                if cell.value == "SUCCESS":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif cell.value == "ERROR":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
                elif cell.value == "WARNING":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C6500")
        
        # Simpan file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"bot_activity_log_{timestamp}.xlsx"
        excel_filepath = os.path.join(BACKUP_PATH, excel_filename)
        
        os.makedirs(BACKUP_PATH, exist_ok=True)
        wb.save(excel_filepath)
        
        logger.info(f"Excel export created: {excel_filepath}")
        return excel_filepath
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        return None

# ============ FUNGSI LOGGING ============
def log_activity(user_info: str, action: str, status: str = "INFO", details: str = ""):
    """Log aktivitas ke file"""
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = {
            "timestamp": timestamp,
            "user": user_info,
            "action": action,
            "status": status,
            "details": details
        }
        
        # Buat folder log jika belum ada
        log_dir = os.path.dirname(LOG_FILE)
        if not os.path.exists(log_dir):
            os.makedirs(log_dir, exist_ok=True)
        
        # Tulis ke file log
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(json.dumps(log_entry, ensure_ascii=False) + "\n")
        
        logger.info(f"Logged: {action} - {user_info}")
        
        # Cek auto backup
        check_auto_backup()
        
    except Exception as e:
        logger.error(f"Error writing log: {e}")

# ============ FUNGSI HELPER ============
async def send_admin_notification(context: ContextTypes.DEFAULT_TYPE, user_info: str, action: str, status: str = "INFO"):
    """Kirim notifikasi ke admin"""
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Tentukan emoji berdasarkan status
        emoji = "ℹ️"
        if status == "SUCCESS":
            emoji = "✅"
        elif status == "ERROR":
            emoji = "❌"
        elif status == "WARNING":
            emoji = "⚠️"
        
        message = f"{emoji} <b>Aktivitas Bot</b>\n\n"
        message += f"👤 User: {user_info}\n"
        message += f"📋 Aksi: {action}\n"
        message += f"⏰ Waktu: {timestamp}"
        
        await context.bot.send_message(
            chat_id=ADMIN_CHAT_ID,
            text=message,
            parse_mode='HTML'
        )
        
        # Log juga ke file
        log_activity(user_info, action, status)
    except Exception as e:
        logger.error(f"Error sending admin notification: {e}")

def get_user_info(update: Update) -> str:
    """Dapatkan informasi user"""
    user = update.effective_user
    username = f"@{user.username}" if user.username else "No username"
    return f"{user.full_name} ({username}) [ID: {user.id}]"

def find_file(person: str, doc_type: str, file_format: str) -> str:
    """Cari file berdasarkan person, document type, dan format"""
    doc_type_clean = doc_type.replace(" ", "_")
    filename = f"{person}_{doc_type_clean}.{file_format.lower()}"
    filepath = os.path.join(BASE_PATH, filename)
    
    # Coba berbagai variasi ekstensi
    variations = [
        filepath,
        filepath.replace(f".{file_format.lower()}", f".{file_format.upper()}"),
        filepath.replace(f".{file_format.lower()}", f".{file_format.capitalize()}")
    ]
    
    for path in variations:
        if os.path.exists(path):
            return path
    
    return None

def get_available_formats(person: str, doc_type: str) -> list:
    """Dapatkan daftar format file yang tersedia untuk person dan doc_type tertentu"""
    available = []
    doc_type_clean = doc_type.replace(" ", "_")
    
    for fmt in FILE_FORMATS:
        filename = f"{person}_{doc_type_clean}.{fmt.lower()}"
        filepath = os.path.join(BASE_PATH, filename)
        
        # Cek berbagai variasi ekstensi
        variations = [
            filepath,
            filepath.replace(f".{fmt.lower()}", f".{fmt.upper()}"),
            filepath.replace(f".{fmt.lower()}", f".{fmt.capitalize()}")
        ]
        
        for path in variations:
            if os.path.exists(path):
                available.append(fmt)
                break
    
    return available

def get_available_documents(person: str) -> list:
    """Dapatkan daftar jenis dokumen yang tersedia untuk person tertentu"""
    available = []
    
    for doc_type in DOCUMENT_TYPES:
        # Cek apakah ada file dengan dokumen ini (format apapun)
        formats = get_available_formats(person, doc_type)
        if formats:  # Jika ada minimal 1 format file
            available.append(doc_type)
    
    return available

async def auto_delete_message(context: ContextTypes.DEFAULT_TYPE, chat_id: int, message_id: int, delay: int):
    """Hapus pesan secara otomatis setelah delay tertentu"""
    try:
        import asyncio
        await asyncio.sleep(delay)
        await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
        logger.info(f"Auto-deleted message {message_id} from chat {chat_id}")
    except Exception as e:
        logger.error(f"Error auto-deleting message: {e}")

# ============ HANDLERS ============
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk command /start"""
    user_id = update.effective_user.id
    user_info = get_user_info(update)
    
    # Cek apakah admin
    is_admin = str(user_id) == str(ADMIN_CHAT_ID)
    
    await send_admin_notification(context, user_info, "Memulai bot dengan /start")
    
    keyboard = []
    for person in PERSONS:
        keyboard.append([InlineKeyboardButton(f"👤 {person}", callback_data=f"person_{person}")])
    
    # Tambah menu admin jika user adalah admin
    if is_admin:
        keyboard.append([InlineKeyboardButton("⚙️ ADMIN MENU", callback_data="admin_menu")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    welcome_text = (
        "📁 <b>Selamat Datang di File Manager Bot</b>\n\n"
        "Pilih nama untuk mengakses dokumen:"
    )
    
    if is_admin:
        welcome_text += "\n\n👑 <i>Anda login sebagai Admin</i>"
    
    await update.message.reply_text(
        welcome_text,
        reply_markup=reply_markup,
        parse_mode='HTML'
    )

async def admin_menu(query, context):
    """Tampilkan menu admin"""
    keyboard = [
        [InlineKeyboardButton("📤 Upload File", callback_data="admin_upload")],
        [InlineKeyboardButton("➕ Tambah Nama", callback_data="admin_addperson")],
        [InlineKeyboardButton("➕ Tambah Jenis Dokumen", callback_data="admin_adddoctype")],
        [InlineKeyboardButton("📋 Lihat Data", callback_data="admin_viewdata")],
        [InlineKeyboardButton("💽 Info Storage", callback_data="admin_storage")],
        [InlineKeyboardButton("📊 Lihat Log", callback_data="admin_viewlog")],
        [InlineKeyboardButton("📥 Export Log (Excel)", callback_data="admin_exportlog")],
        [InlineKeyboardButton("💾 Backup Log Sekarang", callback_data="admin_backuplog")],
        [InlineKeyboardButton("📂 Lihat Backup", callback_data="admin_viewbackup")],
        [InlineKeyboardButton("🗑️ Hapus Log", callback_data="admin_clearlog")],
        [InlineKeyboardButton("◀️ Kembali", callback_data="back_main")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    # Info backup terakhir
    backup_info = ""
    if last_backup_time:
        hours_ago = (datetime.now() - last_backup_time).total_seconds() / 3600
        backup_info = f"\n📅 Backup terakhir: {hours_ago:.1f} jam yang lalu"
    
    # Info ukuran log
    log_size_info = ""
    if os.path.exists(LOG_FILE):
        file_size_mb = os.path.getsize(LOG_FILE) / (1024 * 1024)
        log_size_info = f"\n📊 Ukuran log: {file_size_mb:.2f} MB"
    
    # Info storage yang digunakan
    device = get_mount_device()
    storage_info = f"\n💽 Storage: {device}"
    
    await query.edit_message_text(
        f"⚙️ <b>ADMIN MENU</b>{backup_info}{log_size_info}{storage_info}\n\nPilih menu admin:",
        reply_markup=reply_markup,
        parse_mode='HTML'
    )

async def admin_upload_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mulai proses upload file"""
    query = update.callback_query
    
    keyboard = []
    for person in PERSONS:
        keyboard.append([InlineKeyboardButton(f"👤 {person}", callback_data=f"upload_person_{person}")])
    keyboard.append([InlineKeyboardButton("◀️ Kembali", callback_data="admin_menu")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(
        "📤 <b>Upload File</b>\n\nPilih nama untuk upload dokumen:",
        reply_markup=reply_markup,
        parse_mode='HTML'
    )

async def admin_view_log(query, context):
    """Tampilkan log aktivitas terakhir"""
    try:
        if not os.path.exists(LOG_FILE):
            await query.edit_message_text(
                "📊 <b>Log Aktivitas</b>\n\n"
                "Belum ada log aktivitas.\n\n"
                "Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
            return
        
        # Baca 20 log terakhir
        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        last_logs = lines[-20:] if len(lines) > 20 else lines
        
        log_text = "📊 <b>Log Aktivitas (20 Terakhir)</b>\n\n"
        
        for line in reversed(last_logs):
            try:
                entry = json.loads(line.strip())
                emoji = "ℹ️"
                if entry.get("status") == "SUCCESS":
                    emoji = "✅"
                elif entry.get("status") == "ERROR":
                    emoji = "❌"
                elif entry.get("status") == "WARNING":
                    emoji = "⚠️"
                
                log_text += f"{emoji} <code>{entry['timestamp']}</code>\n"
                log_text += f"   {entry['action']}\n"
                log_text += f"   User: {entry['user'][:30]}...\n\n"
            except:
                continue
        
        # Split jika terlalu panjang
        if len(log_text) > 4000:
            log_text = log_text[:4000] + "\n\n<i>... log dipotong ...</i>"
        
        keyboard = [
            [InlineKeyboardButton("🔄 Refresh", callback_data="admin_viewlog")],
            [InlineKeyboardButton("◀️ Kembali", callback_data="admin_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            log_text,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    except Exception as e:
        logger.error(f"Error viewing log: {e}")
        await query.edit_message_text(
            f"❌ Error membaca log: {str(e)}\n\n"
            "Gunakan /start untuk kembali.",
            parse_mode='HTML'
        )

async def send_approval_request(query, context, person, doc_type, file_format, user_info, user_id):
    """Kirim approval request ke admin"""
    try:
        # Generate unique request ID
        request_id = f"{user_id}_{int(datetime.now().timestamp())}"
        
        # Simpan pending request
        pending_approvals[request_id] = {
            'user_id': user_id,
            'user_info': user_info,
            'person': person,
            'doc_type': doc_type,
            'format': file_format,
            'chat_id': query.message.chat_id,
            'timestamp': datetime.now()
        }
        
        # Kirim notifikasi ke user
        await query.edit_message_text(
            "⏳ <b>Menunggu Persetujuan Admin</b>\n\n"
            f"📄 File: {person} - {doc_type} ({file_format})\n\n"
            "Admin sedang meninjau permintaan Anda.\n"
            "Mohon tunggu sebentar...",
            parse_mode='HTML'
        )
        
        # Cek apakah file ada
        filepath = find_file(person, doc_type, file_format)
        file_status = "✅ Tersedia" if filepath and os.path.exists(filepath) else "❌ Tidak Ada"
        
        # Kirim approval request ke admin
        keyboard = [
            [
                InlineKeyboardButton("✅ Setuju", callback_data=f"approve_{request_id}"),
                InlineKeyboardButton("❌ Tolak", callback_data=f"reject_{request_id}")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        approval_msg = (
            f"📢 <b>PERMINTAAN AKSES FILE</b>\n\n"
            f"👤 User: {user_info}\n"
            f"📁 Person: {person}\n"
            f"📄 Dokumen: {doc_type}\n"
            f"🔎 Format: {file_format}\n"
            f"📊 Status File: {file_status}\n"
            f"⏰ Waktu: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            f"<b>Setujui permintaan ini?</b>"
        )
        
        await context.bot.send_message(
            chat_id=ADMIN_CHAT_ID,
            text=approval_msg,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
        
        log_activity(user_info, f"Request akses: {person} - {doc_type} - {file_format}", "INFO")
        
    except Exception as e:
        logger.error(f"Error sending approval request: {e}")
        await query.edit_message_text(
            "❌ Terjadi kesalahan saat mengirim permintaan.\n\n"
            "Gunakan /start untuk coba lagi."
        )

async def handle_approval(query, context, data, admin_info):
    """Handle approval atau rejection dari admin"""
    try:
        action = "approve" if data.startswith("approve_") else "reject"
        request_id = data.replace(f"{action}_", "")
        
        # Cek apakah request masih ada
        if request_id not in pending_approvals:
            await query.answer("❌ Permintaan sudah tidak valid atau kadaluarsa!", show_alert=True)
            await query.edit_message_text(
                query.message.text + "\n\n❌ <i>Permintaan sudah tidak valid</i>",
                parse_mode='HTML'
            )
            return
        
        request = pending_approvals[request_id]
        user_id = request['user_id']
        user_info = request['user_info']
        person = request['person']
        doc_type = request['doc_type']
        file_format = request['format']
        user_chat_id = request['chat_id']
        
        if action == "approve":
            # Generate password
            password = generate_password()
            user_passwords[user_id] = {
                'password': password,
                'person': person,
                'doc_type': doc_type,
                'format': file_format,
                'timestamp': datetime.now()
            }
            
            # Update approval message
            await query.edit_message_text(
                query.message.text_html + f"\n\n✅ <b>DISETUJUI oleh Admin</b>\n🔑 Password: <code>{password}</code>",
                parse_mode='HTML'
            )
            
            # Kirim password ke admin (pesan terpisah yang bisa di-copy)
            password_msg = await context.bot.send_message(
                chat_id=ADMIN_CHAT_ID,
                text=(
                    f"🔑 <b>PASSWORD AKSES</b>\n\n"
                    f"Password untuk {user_info}:\n"
                    f"<code>{password}</code>\n\n"
                    f"<i>Kirim password ini ke user untuk akses file.</i>"
                ),
                parse_mode='HTML'
            )
            
            # Kirim notifikasi ke user
            await context.bot.send_message(
                chat_id=user_chat_id,
                text=(
                    f"✅ <b>Permintaan Disetujui!</b>\n\n"
                    f"Admin telah menyetujui akses Anda.\n"
                    f"Password akan dikirimkan oleh admin.\n\n"
                    f"Setelah menerima password, silakan masukkan password untuk mengakses file."
                ),
                parse_mode='HTML'
            )
            
            # Set state untuk user
            user_states[user_id] = {
                'awaiting_password': True,
                'request_id': request_id
            }
            
            log_activity(admin_info, f"Approve request dari {user_info} - {person} - {doc_type} - {file_format}", "SUCCESS")
            
        else:  # reject
            # Update approval message
            await query.edit_message_text(
                query.message.text_html + "\n\n❌ <b>DITOLAK oleh Admin</b>",
                parse_mode='HTML'
            )
            
            # Kirim notifikasi ke user
            await context.bot.send_message(
                chat_id=user_chat_id,
                text=(
                    f"❌ <b>Permintaan Ditolak</b>\n\n"
                    f"Admin tidak menyetujui permintaan akses Anda.\n"
                    f"Silakan hubungi admin untuk informasi lebih lanjut.\n\n"
                    f"Gunakan /start untuk kembali ke menu."
                ),
                parse_mode='HTML'
            )
            
            log_activity(admin_info, f"Reject request dari {user_info} - {person} - {doc_type} - {file_format}", "WARNING")
        
        # Hapus pending request
        del pending_approvals[request_id]
        
    except Exception as e:
        logger.error(f"Error handling approval: {e}")
        await query.answer(f"❌ Error: {str(e)}", show_alert=True)

async def admin_clear_log(query, context):
    """Hapus log file"""
    try:
        if os.path.exists(LOG_FILE):
            # Backup dulu sebelum hapus
            backup_path = create_backup()
            
            os.remove(LOG_FILE)
            
            msg = "✅ Log berhasil dihapus!"
            if backup_path:
                msg += f"\n\n💾 Backup tersimpan di:\n<code>{backup_path}</code>"
            
            await query.edit_message_text(
                msg + "\n\nGunakan /start untuk kembali ke menu utama.",
                parse_mode='HTML'
            )
        else:
            await query.edit_message_text(
                "ℹ️ Log sudah kosong.\n\n"
                "Gunakan /start untuk kembali ke menu utama."
            )
    except Exception as e:
        await query.edit_message_text(
            f"❌ Error menghapus log: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def admin_backup_log(query, context):
    """Backup log file sekarang"""
    try:
        status_msg = await query.edit_message_text("⏳ Sedang membuat backup...")
        
        backup_path = create_backup()
        
        if backup_path:
            await status_msg.edit_text(
                f"✅ <b>Backup Berhasil!</b>\n\n"
                f"📁 File: <code>{os.path.basename(backup_path)}</code>\n"
                f"📂 Path: <code>{backup_path}</code>\n\n"
                f"Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
        else:
            await status_msg.edit_text(
                "❌ Gagal membuat backup!\n\n"
                "Gunakan /start untuk kembali."
            )
    except Exception as e:
        await query.edit_message_text(
            f"❌ Error: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def admin_storage_info(query, context):
    """Tampilkan informasi storage detail"""
    try:
        storage = get_storage_info()
        device = get_mount_device()
        
        # Hitung jumlah file
        file_count = 0
        total_size = 0
        if os.path.exists(BASE_PATH):
            for root, dirs, files in os.walk(BASE_PATH):
                for file in files:
                    if not file.startswith('.'):  # Skip hidden files
                        file_count += 1
                        try:
                            filepath = os.path.join(root, file)
                            total_size += os.path.getsize(filepath)
                        except:
                            pass
        
        total_size_mb = total_size / (1024 * 1024)
        
        storage_text = (
            f"💽 <b>INFORMASI STORAGE</b>\n\n"
            f"<b>Device Info:</b>\n"
            f"💿 Device: <code>{storage['device']}</code>\n"
            f"📂 Mount: <code>{device}</code>\n"
            f"📍 Path: <code>{storage['path']}</code>\n\n"
            f"<b>Disk Usage:</b>\n"
            f"💾 Total: {storage['size']}\n"
            f"📊 Terpakai: {storage['used']}\n"
            f"💚 Tersedia: {storage['available']}\n"
            f"📈 Persentase: {storage['use_percent']}\n\n"
            f"<b>Bot Data:</b>\n"
            f"📁 Jumlah file: {file_count}\n"
            f"📦 Total ukuran: {total_size_mb:.2f} MB\n\n"
            f"<i>Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>"
        )
        
        keyboard = [
            [InlineKeyboardButton("🔄 Refresh", callback_data="admin_storage")],
            [InlineKeyboardButton("◀️ Kembali", callback_data="admin_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            storage_text,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    except Exception as e:
        logger.error(f"Error showing storage info: {e}")
        await query.edit_message_text(
            f"❌ Error menampilkan info storage: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def admin_view_data(query, context):
    """Tampilkan data nama dan jenis dokumen"""
    try:
        data_text = "📋 <b>Data Bot</b>\n\n"
        
        data_text += f"👥 <b>Nama ({len(PERSONS)}):</b>\n"
        for i, person in enumerate(PERSONS, 1):
            data_text += f"{i}. {person}\n"
        
        data_text += f"\n📄 <b>Jenis Dokumen ({len(DOCUMENT_TYPES)}):</b>\n"
        for i, doc_type in enumerate(DOCUMENT_TYPES, 1):
            data_text += f"{i}. {doc_type}\n"
        
        data_text += f"\n🔎 <b>Format File ({len(FILE_FORMATS)}):</b>\n"
        data_text += f"{', '.join(FILE_FORMATS)}"
        
        keyboard = [
            [InlineKeyboardButton("🔄 Refresh", callback_data="admin_viewdata")],
            [InlineKeyboardButton("◀️ Kembali", callback_data="admin_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            data_text,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    except Exception as e:
        logger.error(f"Error viewing data: {e}")
        await query.edit_message_text(
            f"❌ Error: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def admin_add_person(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mulai proses tambah nama baru"""
    if update.callback_query:
        query = update.callback_query
        context.user_data['awaiting_new_person'] = True
        
        await query.edit_message_text(
            "➕ <b>Tambah Nama Baru</b>\n\n"
            "Silakan ketik nama baru (huruf kapital).\n"
            "Contoh: BUDI\n\n"
            "<i>Ketik /cancel untuk batal</i>",
            parse_mode='HTML'
        )

async def admin_add_doctype(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mulai proses tambah jenis dokumen baru"""
    if update.callback_query:
        query = update.callback_query
        context.user_data['awaiting_new_doctype'] = True
        
        await query.edit_message_text(
            "➕ <b>Tambah Jenis Dokumen Baru</b>\n\n"
            "Silakan ketik jenis dokumen baru (huruf kapital).\n"
            "Contoh: SERTIFIKAT\n\n"
            "<i>Ketik /cancel untuk batal</i>",
            parse_mode='HTML'
        )

async def admin_export_log(query, context, user_info):
    """Export log ke Excel"""
    try:
        status_msg = await query.edit_message_text("⏳ Sedang mengexport ke Excel...")
        
        excel_path = export_log_to_excel()
        
        if excel_path and os.path.exists(excel_path):
            # Kirim file Excel
            with open(excel_path, 'rb') as f:
                await context.bot.send_document(
                    chat_id=query.message.chat_id,
                    document=f,
                    filename=os.path.basename(excel_path),
                    caption=f"📊 Export Log Bot Activity\n📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                )
            
            await status_msg.edit_text(
                f"✅ <b>Export Berhasil!</b>\n\n"
                f"📁 File Excel telah dikirim.\n"
                f"📂 Tersimpan di: <code>{excel_path}</code>\n\n"
                f"Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
            
            log_activity(user_info, "Export log ke Excel", "SUCCESS")
        else:
            await status_msg.edit_text(
                "❌ Gagal export ke Excel!\n\n"
                "Gunakan /start untuk kembali."
            )
    except Exception as e:
        logger.error(f"Error exporting log: {e}")
        await query.edit_message_text(
            f"❌ Error: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def admin_view_backup(query, context):
    """Tampilkan daftar file backup"""
    try:
        if not os.path.exists(BACKUP_PATH):
            await query.edit_message_text(
                "📂 <b>Daftar Backup</b>\n\n"
                "Belum ada file backup.\n\n"
                "Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
            return
        
        # Ambil semua file backup
        backup_files = []
        for filename in os.listdir(BACKUP_PATH):
            filepath = os.path.join(BACKUP_PATH, filename)
            if os.path.isfile(filepath):
                file_size = os.path.getsize(filepath) / 1024  # KB
                mod_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                backup_files.append({
                    'name': filename,
                    'size': file_size,
                    'time': mod_time
                })
        
        # Sort by time (newest first)
        backup_files.sort(key=lambda x: x['time'], reverse=True)
        
        if not backup_files:
            await query.edit_message_text(
                "📂 <b>Daftar Backup</b>\n\n"
                "Belum ada file backup.\n\n"
                "Gunakan /start untuk kembali.",
                parse_mode='HTML'
            )
            return
        
        # Buat list backup (max 10)
        backup_text = "📂 <b>Daftar Backup</b>\n\n"
        for i, backup in enumerate(backup_files[:10], 1):
            backup_text += f"{i}. <code>{backup['name']}</code>\n"
            backup_text += f"   📊 {backup['size']:.1f} KB | 📅 {backup['time'].strftime('%Y-%m-%d %H:%M')}\n\n"
        
        if len(backup_files) > 10:
            backup_text += f"\n<i>... dan {len(backup_files) - 10} file lainnya</i>\n"
        
        backup_text += f"\n📁 Total: {len(backup_files)} file backup"
        
        keyboard = [
            [InlineKeyboardButton("🔄 Refresh", callback_data="admin_viewbackup")],
            [InlineKeyboardButton("◀️ Kembali", callback_data="admin_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            backup_text,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    except Exception as e:
        logger.error(f"Error viewing backup: {e}")
        await query.edit_message_text(
            f"❌ Error: {str(e)}\n\n"
            "Gunakan /start untuk kembali."
        )

async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk inline keyboard button"""
    query = update.callback_query
    await query.answer()
    
    data = query.data
    user_id = update.effective_user.id
    user_info = get_user_info(update)
    is_admin = str(user_id) == str(ADMIN_CHAT_ID)
    
    # Admin Menu Handlers
    if data == "admin_menu":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_menu(query, context)
        return
    
    elif data == "admin_upload":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_upload_start(update, context)
        return
    
    elif data == "admin_viewlog":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_view_log(query, context)
        return
    
    elif data == "admin_clearlog":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_clear_log(query, context)
        return
    
    elif data == "admin_backuplog":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_backup_log(query, context)
        return
    
    elif data == "admin_exportlog":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_export_log(query, context, user_info)
        return
    
    elif data == "admin_viewbackup":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_view_backup(query, context)
        return
    
    elif data == "admin_viewdata":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_view_data(query, context)
        return
    
    elif data == "admin_addperson":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_add_person(update, context)
        return
    
    elif data == "admin_adddoctype":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_add_doctype(update, context)
        return
    
    elif data == "admin_storage":
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak! Anda bukan admin.")
            return
        await admin_storage_info(query, context)
        return
    
    # Upload handlers
    elif data.startswith("upload_person_"):
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak!")
            return
        
        person = data.replace("upload_person_", "")
        context.user_data['upload_person'] = person
        
        keyboard = []
        for doc_type in DOCUMENT_TYPES:
            keyboard.append([InlineKeyboardButton(
                f"📄 {doc_type}", 
                callback_data=f"upload_doc_{person}_{doc_type}"
            )])
        keyboard.append([InlineKeyboardButton("◀️ Kembali", callback_data="admin_upload")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            f"📤 <b>Upload untuk {person}</b>\n\nPilih jenis dokumen:",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
        return
    
    elif data.startswith("upload_doc_"):
        if not is_admin:
            await query.edit_message_text("❌ Akses ditolak!")
            return
        
        parts = data.replace("upload_doc_", "").split("_", 1)
        person = parts[0]
        doc_type = parts[1]
        
        context.user_data['upload_person'] = person
        context.user_data['upload_doc_type'] = doc_type
        context.user_data['awaiting_file'] = True
        
        await query.edit_message_text(
            f"📤 <b>Upload File</b>\n\n"
            f"👤 Nama: {person}\n"
            f"📄 Dokumen: {doc_type}\n\n"
            f"Silakan kirim file (JPG, PNG, PDF, atau DOCX)\n\n"
            f"<i>File akan disimpan sebagai:\n{person}_{doc_type.replace(' ', '_')}.EXT</i>",
            parse_mode='HTML'
        )
        return
    
    # Handler untuk pilihan person
    elif data.startswith("person_"):
        person = data.replace("person_", "")
        context.user_data['selected_person'] = person
        
        await send_admin_notification(context, user_info, f"Memilih person: {person}")
        
        # Cek dokumen yang tersedia untuk person ini
        available_docs = get_available_documents(person)
        
        if not available_docs:
            await send_admin_notification(
                context, 
                user_info, 
                f"⚠️ Tidak ada dokumen tersedia untuk: {person}",
                "WARNING"
            )
            
            keyboard = [[InlineKeyboardButton("◀️ Kembali", callback_data="back_main")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(
                f"❌ <b>Dokumen Tidak Tersedia</b>\n\n"
                f"Tidak ada dokumen yang tersedia untuk:\n"
                f"👤 {person}\n\n"
                f"Admin telah diberitahu.\n"
                f"Silakan hubungi admin untuk mengupload dokumen.",
                reply_markup=reply_markup,
                parse_mode='HTML'
            )
            return
        
        # Buat keyboard hanya untuk dokumen yang tersedia
        keyboard = []
        for doc_type in available_docs:
            formats_count = len(get_available_formats(person, doc_type))
            emoji = "✅"
            doc_label = f"{emoji} {doc_type}"
            if formats_count > 1:
                doc_label += f" ({formats_count} file)"
            
            keyboard.append([InlineKeyboardButton(
                doc_label, 
                callback_data=f"doc_{person}_{doc_type}"
            )])
        
        keyboard.append([InlineKeyboardButton("◀️ Kembali", callback_data="back_main")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        doc_info = f"\n\n📊 Tersedia: {len(available_docs)} dari {len(DOCUMENT_TYPES)} jenis dokumen"
        
        await query.edit_message_text(
            f"📂 <b>Dokumen untuk {person}</b>\n\n"
            f"Pilih jenis dokumen yang tersedia:{doc_info}",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    
    # Handler untuk pilihan document type
    elif data.startswith("doc_"):
        parts = data.replace("doc_", "").split("_", 1)
        person = parts[0]
        doc_type = parts[1]
        
        context.user_data['selected_doc_type'] = doc_type
        
        await send_admin_notification(context, user_info, f"Memilih dokumen: {person} - {doc_type}")
        
        # Cek file yang tersedia
        available_formats = get_available_formats(person, doc_type)
        
        if not available_formats:
            await send_admin_notification(
                context, 
                user_info, 
                f"⚠️ Tidak ada file tersedia untuk: {person} - {doc_type}",
                "WARNING"
            )
            
            keyboard = [[InlineKeyboardButton("◀️ Kembali", callback_data=f"person_{person}")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(
                f"❌ <b>File Tidak Tersedia</b>\n\n"
                f"Tidak ada file yang tersedia untuk:\n"
                f"👤 {person}\n"
                f"📄 {doc_type}\n\n"
                f"Admin telah diberitahu.\n"
                f"Silakan hubungi admin untuk mengupload file.",
                reply_markup=reply_markup,
                parse_mode='HTML'
            )
            return
        
        # Buat keyboard hanya untuk format yang tersedia
        keyboard = []
        for fmt in available_formats:
            emoji = "✅"
            keyboard.append([InlineKeyboardButton(
                f"{emoji} {fmt}", 
                callback_data=f"fmt_{person}_{doc_type}_{fmt}"
            )])
        
        keyboard.append([InlineKeyboardButton("◀️ Kembali", callback_data=f"person_{person}")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        file_info = f"\n\n📊 Tersedia: {len(available_formats)} dari {len(FILE_FORMATS)} format"
        
        await query.edit_message_text(
            f"📂 <b>{person} - {doc_type}</b>\n\n"
            f"Pilih format file yang tersedia:{file_info}",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    
    # Handler untuk pilihan format file
    elif data.startswith("fmt_"):
        parts = data.replace("fmt_", "").split("_")
        person = parts[0]
        doc_type = "_".join(parts[1:-1])
        file_format = parts[-1]
        
        # Kirim approval request ke admin
        await send_approval_request(query, context, person, doc_type, file_format, user_info, user_id)
    
    # Handler untuk back to main menu
    elif data == "back_main":
        keyboard = []
        for person in PERSONS:
            keyboard.append([InlineKeyboardButton(f"👤 {person}", callback_data=f"person_{person}")])
        
        if is_admin:
            keyboard.append([InlineKeyboardButton("⚙️ ADMIN MENU", callback_data="admin_menu")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            "📁 <b>File Manager Bot</b>\n\nPilih nama untuk mengakses dokumen:",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    
    # Handler untuk approval (approve/reject)
    elif data.startswith("approve_") or data.startswith("reject_"):
        if not is_admin:
            await query.answer("❌ Hanya admin yang bisa approve/reject!", show_alert=True)
            return
        
        await handle_approval(query, context, data, user_info)

async def handle_file_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk upload file dari admin"""
    user_id = update.effective_user.id
    user_info = get_user_info(update)
    
    # Cek apakah admin
    if str(user_id) != str(ADMIN_CHAT_ID):
        await update.message.reply_text("❌ Anda tidak memiliki akses untuk upload file.")
        return
    
    # Cek apakah sedang dalam mode upload
    if not context.user_data.get('awaiting_file', False):
        return
    
    try:
        person = context.user_data.get('upload_person')
        doc_type = context.user_data.get('upload_doc_type')
        
        # Dapatkan file
        file_obj = None
        file_ext = None
        
        if update.message.photo:
            file_obj = update.message.photo[-1]
            file_ext = "JPG"
        elif update.message.document:
            file_obj = update.message.document
            filename = file_obj.file_name
            file_ext = filename.split('.')[-1].upper()
        
        if not file_obj:
            await update.message.reply_text("❌ Format file tidak didukung!")
            return
        
        # Validasi ekstensi
        if file_ext not in FILE_FORMATS:
            await update.message.reply_text(
                f"❌ Format {file_ext} tidak didukung!\n"
                f"Format yang didukung: {', '.join(FILE_FORMATS)}"
            )
            return
        
        # Download file
        status_msg = await update.message.reply_text("⏳ Sedang mengupload file...")
        
        file = await file_obj.get_file()
        
        # Simpan file dengan nama yang sesuai
        doc_type_clean = doc_type.replace(" ", "_")
        filename = f"{person}_{doc_type_clean}.{file_ext.lower()}"
        filepath = os.path.join(BASE_PATH, filename)
        
        # Buat folder jika belum ada
        os.makedirs(BASE_PATH, exist_ok=True)
        
        await file.download_to_drive(filepath)
        
        # Clear upload state
        context.user_data['awaiting_file'] = False
        context.user_data['upload_person'] = None
        context.user_data['upload_doc_type'] = None
        
        await status_msg.edit_text(
            f"✅ <b>File berhasil diupload!</b>\n\n"
            f"📁 Nama file: <code>{filename}</code>\n"
            f"📂 Path: <code>{filepath}</code>\n"
            f"👤 Person: {person}\n"
            f"📄 Dokumen: {doc_type}\n\n"
            f"Gunakan /start untuk kembali ke menu.",
            parse_mode='HTML'
        )
        
        log_activity(
            user_info,
            f"Upload file: {filename}",
            "SUCCESS",
            f"Person: {person}, Doc: {doc_type}"
        )
        
        # Delete file asli yang diupload user (untuk keamanan)
        context.application.create_task(
            auto_delete_message(context, update.message.chat_id, update.message.message_id, 5)
        )
        
    except Exception as e:
        logger.error(f"Error uploading file: {e}")
        await update.message.reply_text(
            f"❌ Error saat upload file: {str(e)}\n\n"
            "Gunakan /start untuk coba lagi."
        )
        
        context.user_data['awaiting_file'] = False

async def handle_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk input password dan input teks lainnya"""
    user_id = update.effective_user.id
    user_info = get_user_info(update)
    is_admin = str(user_id) == str(ADMIN_CHAT_ID)
    
    # Skip jika sedang upload file
    if context.user_data.get('awaiting_file', False):
        return
    
    # Handler untuk tambah nama baru (admin only)
    if is_admin and context.user_data.get('awaiting_new_person', False):
        new_person = update.message.text.strip().upper()
        
        if not new_person:
            await update.message.reply_text("❌ Nama tidak boleh kosong!")
            return
        
        if len(new_person) > 30:
            await update.message.reply_text("❌ Nama terlalu panjang (max 30 karakter)!")
            return
        
        if new_person in PERSONS:
            await update.message.reply_text(f"❌ Nama '{new_person}' sudah ada dalam daftar!")
            return
        
        PERSONS.append(new_person)
        save_data()
        
        context.user_data['awaiting_new_person'] = False
        
        await update.message.reply_text(
            f"✅ <b>Nama berhasil ditambahkan!</b>\n\n"
            f"👤 Nama: {new_person}\n"
            f"📊 Total nama: {len(PERSONS)}\n\n"
            f"Gunakan /start untuk kembali ke menu.",
            parse_mode='HTML'
        )
        
        log_activity(user_info, f"Menambah nama baru: {new_person}", "SUCCESS")
        return
    
    # Handler untuk tambah jenis dokumen baru (admin only)
    if is_admin and context.user_data.get('awaiting_new_doctype', False):
        new_doctype = update.message.text.strip().upper()
        
        if not new_doctype:
            await update.message.reply_text("❌ Jenis dokumen tidak boleh kosong!")
            return
        
        if len(new_doctype) > 50:
            await update.message.reply_text("❌ Nama jenis dokumen terlalu panjang (max 50 karakter)!")
            return
        
        if new_doctype in DOCUMENT_TYPES:
            await update.message.reply_text(f"❌ Jenis dokumen '{new_doctype}' sudah ada dalam daftar!")
            return
        
        DOCUMENT_TYPES.append(new_doctype)
        save_data()
        
        context.user_data['awaiting_new_doctype'] = False
        
        await update.message.reply_text(
            f"✅ <b>Jenis dokumen berhasil ditambahkan!</b>\n\n"
            f"📄 Dokumen: {new_doctype}\n"
            f"📊 Total jenis dokumen: {len(DOCUMENT_TYPES)}\n\n"
            f"Gunakan /start untuk kembali ke menu.",
            parse_mode='HTML'
        )
        
        log_activity(user_info, f"Menambah jenis dokumen baru: {new_doctype}", "SUCCESS")
        return
    
    # Handler untuk password
    if user_id in user_states and user_states[user_id].get('awaiting_password', False):
        password_input = update.message.text.strip()
        
        # Hapus pesan password dari user (untuk keamanan)
        context.application.create_task(
            auto_delete_message(context, update.message.chat_id, update.message.message_id, 2)
        )
        
        # Cek apakah user punya password yang di-generate
        if user_id not in user_passwords:
            await update.message.reply_text(
                "❌ <b>Session Expired</b>\n\n"
                "Password Anda sudah tidak valid.\n"
                "Silakan request akses file lagi dengan /start",
                parse_mode='HTML'
            )
            user_states[user_id]['awaiting_password'] = False
            return
        
        user_pass_data = user_passwords[user_id]
        correct_password = user_pass_data['password']
        
        if password_input == correct_password:
            user_states[user_id]['awaiting_password'] = False
            
            await send_admin_notification(context, user_info, "✅ Password benar - file dikirim", "SUCCESS")
            
            # Ambil data file dari user_passwords
            person = user_pass_data['person']
            doc_type = user_pass_data['doc_type']
            file_format = user_pass_data['format']
            
            filepath = find_file(person, doc_type, file_format)
            
            if filepath and os.path.exists(filepath):
                try:
                    status_msg = await update.message.reply_text("⏳ Password benar! Sedang mengirim file...")
                    
                    if file_format.upper() in ["JPG", "PNG"]:
                        with open(filepath, 'rb') as f:
                            sent_msg = await context.bot.send_photo(
                                chat_id=update.effective_chat.id,
                                photo=f,
                                caption=f"📄 {person} - {doc_type}.{file_format}"
                            )
                    else:
                        with open(filepath, 'rb') as f:
                            sent_msg = await context.bot.send_document(
                                chat_id=update.effective_chat.id,
                                document=f,
                                caption=f"📄 {person} - {doc_type}.{file_format}"
                            )
                    
                    await send_admin_notification(
                        context, 
                        user_info, 
                        f"✅ File terkirim: {person} - {doc_type} - {file_format}",
                        "SUCCESS"
                    )
                    
                    success_msg = await update.message.reply_text(
                        "✅ File berhasil dikirim!\n\n"
                        f"<i>File akan otomatis terhapus dalam {AUTO_DELETE_SECONDS} detik</i>\n\n"
                        "Gunakan /start untuk kembali ke menu utama.",
                        parse_mode='HTML'
                    )
                    
                    # Hapus password setelah digunakan
                    del user_passwords[user_id]
                    
                    # Auto-delete messages
                    context.application.create_task(
                        auto_delete_message(context, update.effective_chat.id, status_msg.message_id, 5)
                    )
                    context.application.create_task(
                        auto_delete_message(context, update.effective_chat.id, sent_msg.message_id, AUTO_DELETE_SECONDS)
                    )
                    context.application.create_task(
                        auto_delete_message(context, update.effective_chat.id, success_msg.message_id, AUTO_DELETE_SECONDS)
                    )
                    
                except Exception as e:
                    logger.error(f"Error sending file: {e}")
                    await update.message.reply_text(
                        "❌ Terjadi kesalahan saat mengirim file."
                    )
            else:
                await send_admin_notification(
                    context, 
                    user_info, 
                    f"⚠️ File tidak ditemukan: {person} - {doc_type} - {file_format}",
                    "WARNING"
                )
                
                await update.message.reply_text(
                    f"❌ File tidak ditemukan!\n\n"
                    f"Admin telah diberitahu tentang file yang hilang.\n\n"
                    "Gunakan /start untuk kembali ke menu utama."
                )
        else:
            await send_admin_notification(context, user_info, "❌ Password salah", "WARNING")
            
            error_msg = await update.message.reply_text(
                "❌ <b>Password Salah!</b>\n\n"
                "Silakan coba lagi atau hubungi admin untuk password baru.",
                parse_mode='HTML'
            )
            
            # Auto-delete error message
            context.application.create_task(
                auto_delete_message(context, update.message.chat_id, error_msg.message_id, 10)
            )

async def cancel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk command /cancel"""
    user_id = update.effective_user.id
    
    # Clear semua state
    context.user_data['awaiting_new_person'] = False
    context.user_data['awaiting_new_doctype'] = False
    context.user_data['awaiting_file'] = False
    
    if user_id in user_states:
        user_states[user_id]['awaiting_password'] = False
    
    await update.message.reply_text(
        "❌ <b>Dibatalkan</b>\n\n"
        "Gunakan /start untuk kembali ke menu utama.",
        parse_mode='HTML'
    )

# ============ MAIN ============
def main():
    """Jalankan bot"""
    # Validasi konfigurasi
    if BOT_TOKEN == "YOUR_BOT_TOKEN_HERE" or ADMIN_CHAT_ID == "YOUR_TELEGRAM_ID_HERE":
        print("❌ ERROR: Harap edit BOT_TOKEN dan ADMIN_CHAT_ID terlebih dahulu!")
        return
    
    # Info BASE_PATH yang terdeteksi
    print(f"\n📁 Detecting storage with documents...")
    print(f"✅ BASE_PATH detected: {BASE_PATH}")
    print(f"💽 Device: {get_mount_device()}")
    
    if not os.path.exists(BASE_PATH):
        print(f"⚠️  WARNING: Folder {BASE_PATH} tidak ditemukan!")
        print(f"Bot akan tetap berjalan, tapi file tidak akan ditemukan.")
    else:
        doc_count = 0
        bot_files = ['bot_activity.log', 'bot_data.json', 'bot_service.log']
        try:
            for file in os.listdir(BASE_PATH):
                if not file.startswith('.') and '_' in file and '.' in file:
                    if file not in bot_files and not os.path.isdir(os.path.join(BASE_PATH, file)):
                        ext = file.split('.')[-1].upper()
                        if ext in ['JPG', 'JPEG', 'PDF', 'PNG', 'DOCX', 'DOC']:
                            doc_count += 1
        except:
            pass
        
        if doc_count > 0:
            print(f"✅ Found {doc_count} document(s)")
        else:
            print(f"⚠️  No documents found (folder exists but empty)")
    
    # Load data dari file
    load_data()
    
    # Buat aplikasi
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Tambahkan handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("cancel", cancel_command))
    application.add_handler(CallbackQueryHandler(button_callback))
    application.add_handler(MessageHandler(
        (filters.Document.ALL | filters.PHOTO) & ~filters.COMMAND, 
        handle_file_upload
    ))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_password))
    
    # Jalankan bot
    print("\n" + "="*50)
    print("✅ Bot berjalan...")
    print("="*50)
    print(f"📁 Base Path: {BASE_PATH}")
    print(f"💽 Device: {get_mount_device()}")
    print(f"📊 Log File: {LOG_FILE}")
    print(f"💾 Backup Path: {BACKUP_PATH}")
    print(f"💽 Data File: {DATA_FILE}")
    print(f"👤 Admin ID: {ADMIN_CHAT_ID}")
    print(f"🗑️ Auto-delete: {AUTO_DELETE_SECONDS} detik")
    print(f"⏰ Auto-backup: Setiap {AUTO_BACKUP_HOURS} jam atau {MAX_LOG_SIZE_MB}MB")
    print(f"\n📋 Data:")
    print(f"   👥 Nama: {len(PERSONS)} orang")
    print(f"   📄 Jenis Dokumen: {len(DOCUMENT_TYPES)} jenis")
    print("="*50 + "\n")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()